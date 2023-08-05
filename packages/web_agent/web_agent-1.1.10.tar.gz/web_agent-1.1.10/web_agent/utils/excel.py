import os
from openpyxl import load_workbook


class ExcelHandle(object):

    def __init__(self):
        pass

    @staticmethod
    def get_ogt_config(excel_file):
        ogt_config = list()
        wb = load_workbook(excel_file)
        sheet = wb.active
        for index in range(2, sheet.max_row+1):
            ip = sheet.cell(index, 1).value
            okg_config = sheet.cell(index, 2).value
            if ip not in ["", None] and okg_config not in ["", None]:
                item = {
                    "ip": ip,
                    "ogt_config": okg_config,
                }
                ogt_config.append(item)
        return ogt_config
