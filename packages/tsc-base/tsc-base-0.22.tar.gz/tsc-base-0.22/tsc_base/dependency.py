# 存在简单依赖第三方包的类和方法
import os
import re
from datetime import datetime
try:
    from .func import dict_to_pair, pair_to_dict, merge_dict, get
except:
    from func import dict_to_pair, pair_to_dict, merge_dict, get


class Excel:
    @staticmethod
    def get_excel_th(th_D, horizontal=True, start_r=0, start_c=0, rectangle='$rectangle', sort_key=None, wild_card='$*v', max_deep=None):
        """递归将dict转换为具有单元格位置坐标的excel表头, 用于 xlwt 等, 行列值从0开始

        Args:
            th_D (dict): 会把所有key拿出来作为表头, 然后value中增加 {rectangle:(开始行,结束行,开始列,结束列)}, 注意会覆盖原始value
            horizontal (bool, optional): 是否横向展开表头
            start_r (int, optional): 表头占据的开始行, 通常xlwt从0开始, openpyxl从1开始
            start_c (int, optional): 表头占据的开始列
            rectangle (str, optional): 作为存储每个单元格的4角坐标的key, 也用于sort_key中的func, 不能与th_D中任意key一样
            sort_key (dict, optional): key:{rectangle:{'f':f(k,v),'r':bool},..}; 用于排序同一个字典中的k, 默认不排序
                f: 函数f输入key下面value中的k和v, 返回排序值, k会按照排序值排序, 例如 lambda t:t[0]
                r (bool, optional): 在 f 存在的情况下, 是否倒序排序, 不存在默认False
            wild_card (str, optional): 用于 sort_key 中匹配key的通配符, 如果在sort_key中找不到key则进入wild_card中的value, 不能与th_D中任意key一样
            max_deep (int, optional): 用于递归确定深度, 不能修改

        Returns:
            c, max_deep: int,int; 总占据列宽, 总占据行数
        """
        if max_deep is None:
            x = list(dict_to_pair(th_D))
            for i in x:
                assert len({rectangle, wild_card} & set(i[0])) == 0, 'th_D 中存在 rectangle/wild_card 标记, 请进行修改:' + str(i)
            max_deep = max([len(i[0]) for i in x])
        sort_key = {} if sort_key is None else sort_key
        if rectangle in sort_key:  # 排序
            par = sorted(th_D.items(), key=sort_key[rectangle]['f'], reverse=bool(sort_key[rectangle].get('r')))
        else:
            par = th_D.items()
        c = 0  # 深度优先得到上层的列宽
        for k, v in par:
            start_c_last = start_c + c  # 随着循环, 列的起始点在变化, 行的起始点不变
            if isinstance(v, dict):  # 随着递归, 行每次向下移动一行
                r = 1  # 下面还要分之则只能占一行
                c += Excel.get_excel_th(v, horizontal, start_r+1, start_c_last, rectangle,
                                        sort_key.get(k) or sort_key.get(wild_card), wild_card, max_deep-1)[0]
            else:
                r = max_deep
                c += 1  # 只占一列
                th_D[k] = v = {}
            v[rectangle] = (start_r, start_r+r-1, start_c_last, start_c+c-1)
            v[rectangle] = v[rectangle] if horizontal else [*v[rectangle][2:], *v[rectangle][:2]]
        return c, max_deep

    @staticmethod
    def get_excel_table(doc_L, ignore_th_f=lambda t: t, td_f=lambda t: t[1], horizontal=True, rectangle='$rectangle', **kw):
        """将多个dict转换为excel表格的单元格坐标, 用于生成excel表格

        Args:
            doc_L (list): [{..},..]; 被转换的dict列表
            ignore_th_f (func, optional): 输入 (key_L,value) 返回 (key_L,value) or None, None表示丢弃key_L这个表头
                用于修剪层次过深的表头, 这时候可能需要 td_f 优化dict在单元格中的展示形式
            td_f (func, optional): 用于优化单元格中的值, value是func, 输入 (key_L,value) 返回优化的展示 value
                小心处理每一种值格式, 因为 value 不能是 dict,list,tuple,set 等类型, 否则可能导致 excel 写入出错
            horizontal (bool, optional): 见 get_excel_th
            rectangle (str, optional): 见 get_excel_th
            **kw: 其他参数见 get_excel_th

        Returns:
            dict, list, list: th_D 的一个例子:
                {'inf_jcr': {'$rectangle': (0, 0, 1, 33),
                'Open Access': {'$rectangle': (1, 5, 29, 29)},
                '期刊分区': {'$rectangle': (1, 1, 1, 27),
                        'JCR分区': {'$rectangle': (2, 5, 27, 27)},}}
        """
        # 表头
        pair = []
        for p in dict_to_pair(merge_dict(doc_L)):
            p = ignore_th_f(p)
            if p is not None:
                pair.append(p)
        th_D = pair_to_dict(pair)
        Excel.get_excel_th(th_D, horizontal=horizontal, rectangle=rectangle, **kw)
        th_L = [(i[1], i[0][-2]) for i in dict_to_pair(th_D)]  # 与 td_L 格式一致
        # 表中值
        td_L = []  # [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 与doc_L顺序一致
        for no, doc in enumerate(doc_L):
            coor_v_D = {}  # {(开始行,结束行,开始列,结束列):单元格值,..}
            for p in dict_to_pair(doc):
                p = ignore_th_f(p)  # (key_L,value); 用于获取行列位置, 位置靠前了值也会变
                # 检查这个值是否满足要求
                if p is None:
                    continue
                th = get(p[0], th_D)  # 对应的表头坐标
                if len(th) > 1:  # 没有到叶子结点, 说明doc存在大段的空
                    continue
                # 获取坐标
                _, r, _, c = th[rectangle]  # 右下角单元格坐标
                r, c = (r + no + 1, c) if horizontal else (r, c + no + 1)
                coor = (r, r, c, c)  # (开始行,结束行,开始列,结束列)
                # 保存坐标与单元格值
                if coor not in coor_v_D:  # 防止重复
                    coor_v_D[coor] = td_f((p[0], get(p[0], doc)))
            td_L.append(list(coor_v_D.items()))
        return th_D, th_L, td_L

    @staticmethod
    def excel_add_sheet(workbook, name, th_L, td_L, package, index=0, save_path=None, tqdm_f=None):
        """使用行列坐标 写入一页 excel 表格

        Args:
            workbook (obj): openpyxl.Workbook() or xlwt.Workbook(encoding='utf8')
            name (str): sheet 名称
            th_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            td_L (list): [[((开始行,结束行,开始列,结束列),单元格值),..],..]; 行列编号从0开始, 0行就是第一行
            package : (xlwt>=1.3.0 import xlwt) or (openpyxl>=3.0.9 import openpyxl)
            index (int, optional): 只用于 openpyxl, 表示插入的 sheet 的位置, xlwt 只是追加
            save_path (str, optional): 写入文件的路径, 会自动添加后缀名
            tqdm_f : tqdm>=4.62.3 from tqdm import tqdm

        Returns:
            workbook
        """
        if 'openpyxl' in str(type(workbook)):  # openpyxl - xlsx (打开效率更高)
            worksheet = workbook.create_sheet(name, index)
            # 写表头
            for coor, v in th_L:
                coor = [i+1 for i in coor]
                worksheet.merge_cells(start_row=coor[0], start_column=coor[2], end_row=coor[1], end_column=coor[3])
                cell = worksheet.cell(coor[0], coor[2])
                cell.value = v
                cell.alignment = package.styles.Alignment(horizontal='center', vertical='center')
                cell.font = package.styles.Font(bold=True)
            # 单元格值
            par = tqdm_f(td_L, f'{name}-写入表格(openpyxl)') if tqdm_f else td_L
            for i in par:
                for coor, v in i:
                    cell = worksheet.cell(coor[0]+1, coor[2]+1)
                    cell.value = v
                    cell.alignment = package.styles.Alignment(horizontal='center', vertical='center')
            save_path = save_path + '.xlsx' if save_path else None
        else:  # xlwt - xls (写入快6,7倍)
            worksheet = workbook.add_sheet(name)
            # 单元格值
            style = package.XFStyle()
            alignment = package.Alignment()
            alignment.horz = package.Alignment.HORZ_CENTER
            alignment.vert = package.Alignment.VERT_CENTER
            style.alignment = alignment
            par = tqdm_f(td_L, f'{name}-写入表格(xlwt)') if tqdm_f else td_L
            for i in par:
                for coor, v in i:
                    worksheet.write_merge(*coor, v, style)
            # 写表头
            font = package.Font()
            font.bold = True
            style.font = font
            for coor, v in th_L:
                worksheet.write_merge(*coor, v, style)
            save_path = save_path + '.xls' if save_path else None
        if save_path:
            if not os.path.exists(os.path.dirname(save_path)):
                os.makedirs(os.path.dirname(save_path))
            workbook.save(save_path)
        return worksheet


def range_time(start_t, end_t, delta={'weekday': 0, 'weeks': 1}, to_timestamp=True):
    """将时间划分为一个个的间隔, 类似于 range 递进分割时间

    Args:
        start_t (datetime.datetime, int, float): 开始时间, int or float 表示时间戳
        end_t (datetime.datetime, int, float): 结束时间
        delta (dict, optional): 从早到晚按 delta 分割, 所以最后一个时间段可能划分不满
            {'weekday': 0, 'weeks': 1} 表示一周一个间隔, 但是第2个间隔从下周一开始
            内部参数参见: https://dateutil.readthedocs.io/en/stable/relativedelta.html
        to_timestamp (bool, optional): 是否输出成 float 的时间戳格式, 否则是 datetime.datetime

    Returns:
        ([float,..] or [datetime.datetime,..]): interval_L, 第一个和最后一个值必然是 start_t 和 end_t
    """
    from dateutil import parser, relativedelta  # python-dateutil>=2.8.2
    # str to datetime
    start_t = datetime.timestamp(parser.parse(start_t)) if isinstance(start_t, str) else start_t
    end_t = datetime.timestamp(parser.parse(end_t)) if isinstance(end_t, str) else end_t
    # timestamp to datetime
    start_t = datetime.fromtimestamp(start_t) if type(start_t) in {int, float} else start_t
    end_t = datetime.fromtimestamp(end_t) if type(end_t) in {int, float} else end_t
    assert end_t >= start_t, f'end_t < start_t: {end_t} < {start_t}'
    # 第一个间隔计算
    interval_L = [start_t]
    s = start_t + relativedelta.relativedelta(**delta) - relativedelta.relativedelta(**delta)
    if s != interval_L[-1]:
        interval_L.append(s)
    # 后面所有间隔计算
    is_end = interval_L[-1] >= end_t
    while not is_end:
        interval_L.append(interval_L[-1] + relativedelta.relativedelta(**delta))
        assert interval_L[-1] != interval_L[-2], f'delta 没有增加时间: {interval_L[-1]} + {str(delta)}'
        is_end = interval_L[-1] >= end_t
    interval_L[-1] = end_t
    # 转时间戳
    interval_L = [datetime.timestamp(t) for t in interval_L] if to_timestamp else interval_L
    return interval_L


class Bib:
    @staticmethod
    def get_bib(tex_path_L, bib_path_L, bibtexparser):
        """获取所有bib_path_L中的参考文献,并且ID在所有tex_path_L中出现过的

        Args:
            tex_path_L (list): [str,..]; 所有tex文件路径
            bib_path_L (list): [str,..]; 所有bib文件路径
            bibtexparser : bibtexparser>=1.2.0 import bibtexparser

        Returns:
            list: [{'title':..,..},..]
        """
        # 获取所有tex文件,用于匹配bib文献是否在论文中出现过. 如果被注释或者ID是子串可能导致匹配过多
        tex = ''
        for tex_path in tex_path_L:
            with open(tex_path, 'r', encoding='utf8') as r:
                tex += r.read() + '\n'
        # 匹配bib
        bib_entries = []
        for bib_path in bib_path_L:  # 因为可能来自不同的bib, 所以id重复不代表论文重复
            with open(bib_path, 'r', encoding='utf8') as r:
                # 防止出现 bibtexparser.bibdatabase.UndefinedString: 'jan' 错误: https://github.com/sciunto-org/python-bibtexparser/issues/222#issuecomment-427624923
                parser = bibtexparser.bparser.BibTexParser(common_strings=True)
                for bib in bibtexparser.load(r, parser).entries:
                    if re.search('\\\\[^\s\\{}]+{[^{}]*?' + re.escape(bib['ID']) + '[^{}]*?}', tex):  # 不是cite有可能误操作
                        bib_entries.append(bib)
                    else:
                        continue
        return bib_entries

    @staticmethod
    def get_tex_bib_path(dir, tex_blacklist=None):
        """递归获取一个目录中的所有tex和bib文件

        Args:
            dir (str): 主目录
            tex_blacklist (set, list, optional): 不包含的tex文件名

        Returns:
            tex_path_L, bib_path_L: [str,..],[str,..]
        """
        tex_blacklist = set(tex_blacklist) if tex_blacklist else set()
        tex_path_L, bib_path_L = [], []
        for root, _, files in os.walk(dir):
            for name in files:
                if name in tex_blacklist:
                    continue
                if name[0] not in {'.', '_'}:
                    if name[-4:] == '.tex':
                        tex_path_L.append(os.path.join(root, name))
                    if name[-4:] == '.bib':
                        bib_path_L.append(os.path.join(root, name))
        return tex_path_L, bib_path_L

    @staticmethod
    def bib_check(bib_entries, ref_list_path=None, out_duplicate=True):
        """检查提取bib参考文献是否和论文一样,是否有重复等问题

        Args:
            bib_entries (list): [{'title':..,..},..]; bib列表
            ref_list_path (str, optional): 从论文中拷贝的参考文献文件路径, 用于分析bib中的文献是否都是论文中出现的
            out_duplicate (bool, optional): 是否输出重复论文的信息

        Returns:
            bib_entries_: 去重复标题后的 bib_entries, 标题会全小写并去字母以外字符后匹配
        """
        if ref_list_path:
            with open(ref_list_path, 'r', encoding='utf8') as r:
                ref = r.read().lower().replace('ﬁ', 'fi')
                ref = ref.replace('ﬂ', 'fl')
                ref = ref.replace('ﬀ', 'ff')
                ref = ref.replace('ﬃ', 'ffi')
                ref = re.sub('[^a-z]+', '', ref).strip()
            n = 1
            for bib in bib_entries:
                title = re.sub('[^a-z]+', '', bib['title'].lower())
                if title not in ref:
                    if n == 1:
                        print('bib中检测存在的论文,而复制的pdf文献列表中没有该论文标题 (如果很短的标题属于另一个标题的子标题会识别为存在):')
                    print(n, (bib['ID'], bib['title']))
                    n += 1
            assert n == 1, '这种情况需要检查原因,比如因为加入了不该加入的tex文件导致'
        # 输出标题一样的论文
        title_ = {}
        n = 1
        bib_entries_ = []  # 格式与 bib_entries 一致, 去除标题可能一样的bib
        for bib in bib_entries:
            if 'title' not in bib:
                print('没有标题的bib:', bib)
                continue
            title = re.sub('[^a-z]+', '', bib['title'].lower())
            if title in title_:
                if out_duplicate:
                    if n == 1:
                        print('标题可能一样的论文:')
                    print(n, (bib['ID'], bib['title']))
                    print(n, title_[title])
                n += 1
            else:
                title_[title] = (bib['ID'], bib['title'])
                bib_entries_.append(bib)
        print(f"bib论文总数: {len(bib_entries)}, 去标题重复后数量: {len(bib_entries_)}")
        return bib_entries_

    simplify_venue_converter = {
        # 会议
        'NeurIPS': ['Advances in Neural Information Processing Systems'],
        'AAAI': ['.*AAAI.*'],
        'ACL': ['.*Annual Meeting of the Association for Computational Linguistics.*'],
        'UAI': ['.*Conference on Uncertainty in Artificial Intelligence.*'],
        'WWW': ['The World Wide Web Conference', '.*WWW.*'],
        'EMNLP': ['.*Conference on Empirical Methods in Natural Language Processing'],
        'WSDM': ['.*ACM International Conference on Web Search and Data Mining'],
        'CVPR': ['.*Conference on Computer Vision and Pattern Recognition', '.*CVPR.*'],
        'SIGKDD': ['.*SIGKDD.*'],
        'ECCV': ['.*European conference on computer vision.*'],
        'ICCV': ['.*ICCV.*', '.*international conference on computer vision.*'],
        'ICML': ['.*International Conference on Machine Learning.*', '.*ICML.*'],
        'AISTATS': ['.*Artificial Intelligence and Statistics.*'],
        'IJCAI': ['.*International Joint Conference on Artificial Intelligence.*'],
        'CIKM': ['.*Conference on Information and Knowledge Management.*', '.*CIKM.*'],
        'ICLR': ['.*International Conference on Learning Representations.*', '.*ICLR.*'],
        'NAACL': ['.*North American Chapter of the Association for Computational Linguistics.*'],
        'ICDM': ['.*International Conference on Data Mining.*'],
        'WACV': ['.*WACV.*'],
        'DASFAA': ['.*Database Systems for Advanced Applications.*'],
        'PACMCGIT': ['.*Computer graphics and interactive techniques.*'],
        # 期刊
        'arXiv': ['arXiv.*'],
        'CoRR': ['CoRR.*'],
        'TKDE': ['.*TKDE.*'],
        'TNNLS': ['.*TNNLS.*'],
        'TCYB': ['.*TCYB.*'],
        'JMLR': ['.*Journal of Machine Learning Research.*'],
        'KBS': ['Knowledge-Based Systems'],
        'ToG': ['.*Transactions on Graphics.*'],
        'TACL': ['.*Transactions of the Association for Computational Linguistics.*'],
        'PNAS': ['.*Proceedings of the National Academy of Sciences.*'],
        'TPAMI': ['IEEE Trans.+? Pattern Anal.+? Mach.+? Intell.+?', '.*TPAMI.*'],
    }

    @staticmethod
    def simplify_venue(venue, converter_D=None):
        '''
        简化venue的会议/期刊/书籍, 并缩写
        :param venue: str
        :param converter_D: {简称:[匹配,..]}; 匹配是str表示用不区分大小写(全大写还是会区分)的re.search,是func就输入venuc输出是否匹配
        :return: venue
        '''
        converter_D = converter_D if converter_D else Bib.simplify_venue_converter
        venue = re.sub('\s+', ' ', venue).strip()
        for shorthand, converter_L in converter_D.items():
            for converter in converter_L:
                if isinstance(converter, str):
                    if converter.upper() == converter:
                        flags = 0
                    else:
                        flags = re.IGNORECASE
                    if re.search(converter, venue, flags):
                        venue = shorthand
                        break
                else:
                    if converter(venue):
                        venue = shorthand
                        break
        return venue

    @staticmethod
    def simplify_bib_venue(bib_entries, converter_D=None, out_no=True, venue='Venue'):
        """简化所有bib字典中的venue

        Args:
            bib_entries (list): [{'title':..,..},..]; bib列表, 会直接在这个变量上修改
            converter_D (dict, optional): {简称:[匹配,..]}; 匹配是str表示用不区分大小写(全大写还是会区分)的re.search,是func就输入venuc输出是否匹配
            out_no (bool, optional): 是否输出 没有venue的bib,没有被简化的venue
            venue (str, optional): bib增加的字段,用于保存简化后的结果(没有简化也会复制一遍到这)

        Returns:
            list: mark_L, 与bib_entries顺序对应, 描述这条记录的mark符号
        """
        converter_D = converter_D if converter_D else Bib.simplify_venue_converter
        mark_L = []
        venue_field_S = set()
        for bib in bib_entries:
            entryType = bib['ENTRYTYPE']
            if entryType == 'misc':
                venue_field = 'archiveprefix'
                mark = '[J].'
            elif entryType == 'article':
                venue_field = 'journal'
                mark = '[J].'
            elif entryType == 'techreport':
                venue_field = 'institution'
                mark = '[J].'
            elif entryType == 'inproceedings':
                venue_field = 'booktitle'
                mark = '[C]//'
            elif entryType == 'incollection':
                venue_field = 'booktitle'
                mark = '[C]//'
            elif entryType == 'book':
                venue_field = 'publisher'
                mark = '[M].'
            elif entryType == 'inbook':
                venue_field = 'booktitle'
                mark = '[M].'
            elif entryType == 'phdthesis':
                venue_field = 'university'
                mark = '[D].'
            else:  # 未知类型
                raise NameError('未知文献类型! ' + str(bib))
            mark_L.append(mark)
            if venue_field in bib:
                bib[venue] = Bib.simplify_venue(bib[venue_field], converter_D)
                venue_field_S.add(bib[venue])
            elif out_no:
                print('没有venue:', bib)
        if out_no:
            no_simplify = []
            for i in venue_field_S - set(converter_D):
                if i.upper().replace(' ', '') == i:  # 全大写缩写不考虑了
                    continue
                no_simplify.append(i)
            print(f'没有被简化的venue({len(no_simplify)}):')
            for i in no_simplify:
                print(i)
        return mark_L


if __name__ == '__main__':
    # Excel
    print('=' * 10, 'Excel')
    doc_L = [{
        'a1': {'b1': 1, 'b2': 2, 'b3': {'c1': 1, 'c2': 2}},
        'a2': 123,
    } for i in range(10)]
    th_D, th_L, td_L = Excel.get_excel_table(doc_L)
    print('th_D:', th_D)
    import xlwt
    import openpyxl
    from tqdm import tqdm
    Excel.excel_add_sheet(openpyxl.Workbook(), 'test', th_L, td_L, openpyxl, 0, 'test/openpyxl', tqdm)
    Excel.excel_add_sheet(xlwt.Workbook(encoding='utf8'), 'test', th_L, td_L, xlwt, 0, 'test/xlwt', tqdm)

    # range_time
    print('=' * 10, 'range_time')
    interval_L = range_time('2022-02-03', '2022-03-03', to_timestamp=False)
    for i in interval_L:
        print(i)

    # Bib
    print('=' * 10, 'Bib')
    bibs = """
    @incollection{GM05,
    Author= "Dan Geiger and Christopher Meek",
    Title= "Structured Variational Inference Procedures and their Realizations (as incol)",
    Year= 2005,
    Booktitle="Proceedings of Tenth International Workshop on Artificial Intelligence and Statistics, {\rm The Barbados}",
    Publisher="The Society for Artificial Intelligence and Statistics",
    Month= jan,
    Editors= "Z. Ghahramani and R. Cowell"
    }
    """
    with open('test/ref.bib', 'w', encoding='utf8') as w:
        w.write(bibs)
    with open('test/ref.tex', 'w', encoding='utf8') as w:
        w.write('\\citep{GM05}')
    tex_path_L, bib_path_L = Bib.get_tex_bib_path('test')
    import bibtexparser
    bib_entries = Bib.get_bib(tex_path_L, bib_path_L, bibtexparser)
    print('mark_L:', Bib.simplify_bib_venue(bib_entries))
    print('bib_entries:', Bib.bib_check(bib_entries))
