"""firebehaviour.py

Defines the classes used to analyse fire behaviour with PyroPy.

"""

from dataclasses import dataclass
import warnings
from openpyxl import load_workbook
from pandas import DataFrame, read_csv

if __name__ == '__main__':
    import spreadmodels as fbs
    import helpers as fbh
else:
    from . import spreadmodels as fbs
    from . import helpers as fbh

class Incident(object):
    """A wildfire incident.

    Attributes:
        df (Dataframe): Weather and model output data
        wrf (float): wind reduction factor (0-6)
        fuel_load (float): fine fuel load (t/ha)
        fhs_surf (float): surface fuels hazard score (1-4)
        fhs_n_surf (float): near surface fuels hazard score (1-4)
        fuel_height_ns (float): near surface fuel height (cm)

    """
    def __init__(self, weather_df: DataFrame):
        self.df = weather_df
        self.wrf = None
        self.fuel_load = None #t/ha
        self.fhs_surf = None
        self.fhs_n_surf = None 
        self.fuel_height_ns = None

        self.fbcalc_params_mk5 = {
            'wrf': 'J4',
            'fuel_load': 'C4', 
        }

        self.fbcalc_params_vesta = {
            'fhs_surf': 'D6',
            'fhs_n_surf': 'D7',
            'fuel_height_ns': 'C8',
        }

    
    def get_params(self) -> dict:
        """Gets the model parameters that have been defined.

        Returns:
            dict: dictionary with model parameter names and values
        """
        params = list(self.__dict__.items())
        params = dict(params[1:]) #drop the dataframe
        params = {key: val for key, val in params.items() if isinstance(val,(int,float))}
        return params

    def get_df(self) -> DataFrame:
        """

        Returns:
            DataFrame: the Incident Data as a pandas `Dataframe`
        """
        return self.df

    def get_models(self) -> list:
        """

        Returns:
            list: a list of the models that have been run for the Incident
        """
        models = {
            'forest_mk5': 'fros_mk5',
            'forest_vesta': 'fros_vesta',
        }

        return [
            key for key, val in models.items() 
            if val in self.df.columns.values
        ]

    def set_params(self, params: dict) -> None:
        """Update several model parameters using a dictionary.

        The dictionary keys must match the name of the parameter.

        Args:
            params (dict): a dictionary of the model parameters to be updated.
        """
        for key, val in params.items():
            setattr(self, key, val)

    def run_forest_mk5(self) -> None:
        """Runs the McArthur Mk5 Forest Fire Danger Meter model.

        Adds the results to the `Incident.df`
        """
        forest_mk5_params = {
            'wrf': self.wrf,
            'fuel_load': self.fuel_load,
        }
        if self.check_params(forest_mk5_params):
            self.df = fbs.ros_forest_mk5(self.df, self.wrf, self.fuel_load)

    def run_forest_vesta(self) -> None:
        """Runs the Project Vesta (fuel hazard scores) model.

        Adds the results to the `Incident.df`
        """
        forest_vesta_params = {
            'fhs_surf': self.fhs_surf,
            'fhs_n_surf': self.fhs_n_surf,
            'fuel_height_ns': self.fuel_height_ns,
        }


        if self.check_params(forest_vesta_params):
            self.df = fbs.ros_forest_vesta(
                self.df, 
                self.fhs_surf, 
                self.fhs_n_surf, 
                self.fuel_height_ns,
            )

    def print(self, head=False) -> None:
        """Prints the field headings and rows of the `Dataframe`

        Args:
            head (bool, optional): Print only the head (first 5 rows).
                Defaults to False.
        """
        if head: print(self.df.head())
        else: print(self.df)

    def check_params(self, params: dict) -> bool:
        """Checks to see if parameters have been defined.

        Args:
            params (dict): a dictionary with the parmeters to check.

        Returns:
            bool: `True` is values for the parameters have been defined,
                else `False`
        """
        incident_params = self.get_params()
        for key in params.keys():
            if not key in incident_params.keys():
                warnings.warn(f'{key} not set - run set params')
                return False
        return True

    def compare_fbcalc(self, fn: str, models: list) -> None:
        """Loads results from an FireBehaviourCalcs spreadsheet into the 
        `Incident.df`.

        Only loads the pages from FireBehaviourCalcs that have correspond 
        to models already in the `Incident.df` 

        Args:
            fn (str): path to the FireBehaviourCalcs spreadsheet
            model (list): a list of the models to compare. Valid items
                include: `'mk5'` McArthur mk5 model, `'vesta'` The Project 
                VESTA model. Also there is a little easter egg here and if you 
                put 'mc' as the model it will get the moisture content (%)
                from the VESTA model.

        Returns:
            None:
        """
        fbcalc_refs = {
            'mk5': ['Forest(McArthur)', 'O', self.fbcalc_params_mk5],
            'vesta': ['Forest(VESTA)', 'P', self.fbcalc_params_vesta],
            'mc': ['Forest(VESTA)', 'M', {}],
        }

        run_model_functions = {
            'mk5': self.run_forest_mk5,
            'vesta': self.run_forest_vesta,
        }

        if fbh.check_filepath(fn, suffix='xlsm'):
            wb = load_workbook(fn, data_only=True) #, keep_vba=True
            for model, ref in fbcalc_refs.items():
                if model in models:
                    sheet_name, column, model_params = ref
                    ws = wb[sheet_name]
                    column = ws[column]
                    ros_vals =  [cell.value for cell in column if isinstance(
                            cell.value, (float, int)
                        )]
                    self.df[f'{model}_fbcalc'] = ros_vals
                    self.df[f'{model}_fbcalc'] = self.df[f'{model}_fbcalc'].astype(int)
                    params = {param: ws[address].value 
                        for (param, address) in model_params.items()
                    }

                    self.set_params(params)
                    if model in run_model_functions.keys(): #trick it into getting columns that aren't models like mc
                        run_model_functions[model]()            
        return None

    def compare_amicus(self, fn: str) -> None:
        if fbh.check_filepath(fn, suffix='csv'):
            print(fbh.check_encoding(fn))
            df = read_csv(fn, header=0, encoding=fbh.check_encoding(fn))
            self.df['mc_amicus'] = df['Predicted FMC (%)']
            self.df['vesta_amicus'] = df['Rate of spread (m/h)']
        pass

    def set_fbcalc(self, fn: str) -> bool:
        """Writes the weather data and model parameters to a FireBehaviourCalc
        macro enabled spreadsheet.

        ** WARNING! THE VALUES IN THE SPREADSHEET WILL BE OVERWRITTEN **

        Spreadsheet macros will not run (ie changes will not take place in 
        the spreadsheet) until you open it and click on the relevant sheets.

        Args:
            fn (str): path to the FireBehaviourCalc spreadsheet. 
                 
        
        Returns:
            `True` f successful else `False`
        """

        models = {
            'forest_mk5': ['Forest(McArthur)', self.fbcalc_params_mk5],
            'forest_vesta': ['Forest(VESTA)', self.fbcalc_params_vesta],
        }

        if fbh.check_filepath(fn, suffix='xlsm'):
            datetime_format: str = "%d/%m/%Y %H:%M"
            # weather = self.df.iloc[:, 0:6]
            weather = self.df.copy(deep=True)
            weather['date'] = weather['date_time'].dt.strftime(datetime_format.split()[0])
            weather['time'] = weather['date_time'].dt.strftime(datetime_format.split()[1])
            # re-order and remove unwanted cols
            weather_cols = ['date', 'time', 'temp', 'humidity', 'wind_dir', 'wind_speed', 'drought']
            weather = weather[weather_cols]

            wb = load_workbook(fn, read_only=False, keep_vba=True)

            #load weather data to fbcalc
            ws = wb['Weather_Site']
            row_ctr = 0
            ws_weather_cols = [2,8]
            ws_weather_rows = [12, 12+weather.shape[0]-1]
            for row in ws.iter_rows(
                min_col=ws_weather_cols[0],
                max_col=ws_weather_cols[1],
                min_row=ws_weather_rows[0],
                max_row=ws_weather_rows[1],
            ):
                cell_ctr = 0
                for cell in row:
                    cell.value = weather.iloc[row_ctr, cell_ctr]
                    cell_ctr += 1
                row_ctr += 1

            #load parameters to fbcalc
            for model in self.get_models():
                ws = wb[models[model][0]]
                for param, addr in models[model][1].items():
                    ws[addr] = self.__dict__[param]

            return True

        return False            



if __name__ == '__main__':
    pass