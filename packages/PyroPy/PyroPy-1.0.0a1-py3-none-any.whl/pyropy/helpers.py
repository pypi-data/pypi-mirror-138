"""Utility scripts for fire behaviour analysis."""
import os
import warnings
from openpyxl import Workbook, load_workbook

if  __name__ == '__main__':
    from firebehaviour import Incident
else:
    from .firebehaviour import Incident

def check_filepath(fn: str, suffix: str = None) -> bool:
    valid = os.path.isfile(fn)
    if not valid: warnings.warn(f'{fn} is not a valid filename')
    if suffix:
        fn = fn.split('.')
        valid = (fn[-1] == suffix)
        if not valid: warnings.warn(f'file must be a *.{suffix} file')
    return valid

def check_encoding(fn: str) -> str:
    """Returns the encoding of a csv file.
    
    TODO if encoding is Windws cp1252 change to latin-1 for better mapping see 
    http://python-notes.curiousefficiency.org/en/latest/python3/text_file_processing.html
    """
    with open(fn) as f:
        return(f.encoding)

def incident_to_calc(
        incident: Incident,
        calc_fn: str, 
        datetime_format: str = "%d/%m/%Y %H:%M",
    ) -> Workbook:
    """reads incident data and parameters into fb_calc spreadsheet
    """
    weather = incident.df.iloc[:, 0:6]
    weather['date'] = weather['date_time'].dt.strftime(datetime_format.split()[0])
    weather['time'] = weather['date_time'].dt.strftime(datetime_format.split()[1])
    # re-order and remove unwanted cols
    weather_cols = ['date', 'time', 'temp', 'humidity', 'wind_dir', 'wind_speed', 'drought']
    weather = weather[weather_cols]

    check_filepath(calc_fn, suffix='xlsm')
    wb = load_workbook(calc_fn, read_only=False, keep_vba=True)

    #load weather data
    ws = wb['Weather_Site']
    row_ctr = 0
    ws_weather_cols = [2,8]
    ws_weather_rows = [12, 12+weather.shape[0]-1]
    for row in ws.iter_rows(
        min_col=ws_weather_cols[0],
        max_col=ws_weather_cols[1],
        min_row=ws_weather_rows[0],
        max_row=ws_weather_rows[1],
        ):
        cell_ctr = 0
        for cell in row:
            cell.value = weather.iloc[row_ctr, cell_ctr]
            cell_ctr += 1
        row_ctr += 1
    
    #load model params
    ## forest mk5
    ws = wb['Forest(McArthur)']
    ws['C4'] = incident.fuel_load
    ws['J4'] = incident.wrf
    ## forest vesta
    ws = wb['Forest(VESTA)']
    #TODO what a pain Kevin uses ratings

    ws['D6'] = incident.fhs_surf
    ws['D7'] = incident.fhs_n_surf
    ws['C8'] = incident.fuel_height_ns
    return wb


if __name__ == '__main__':
    pass
